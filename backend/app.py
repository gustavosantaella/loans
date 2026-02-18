import os
import io
import logging
import sqlite3
import tempfile
from datetime import datetime, timedelta, timezone
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from dotenv import load_dotenv
from pymongo import MongoClient
import jwt
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Load environment variables
load_dotenv()

app = Flask(__name__)
CORS(app, origins=[
    'http://localhost:4200',
    'http://127.0.0.1:4200',
    'https://loans-seven.vercel.app',
    'https://loans-front.vercel.app'
])

# Logging configuration
logging.basicConfig(level=logging.INFO)
logger = app.logger

# --- MongoDB ---
MONGO_URI = os.getenv('MONGO_DB_URI')
mongo_client = MongoClient(MONGO_URI)
mongo_db = mongo_client.get_database('loans_db')
logger.info("MongoDB connected successfully to loans_db")

# Collections
col_clients = mongo_db['clients']
col_partners = mongo_db['partners']
col_loans = mongo_db['loans']
col_payments = mongo_db['payments']
col_counters = mongo_db['counters']

# --- Auth Config ---
ADMIN_EMAIL = os.getenv('USER_ADMIN')
ADMIN_PASSWORD = os.getenv('USER_ADMIN_PASSWORD')
JWT_SECRET = os.getenv('JWT_SECRET', 'fallback_secret')


# ===================== Auth =====================

@app.route('/api/login', methods=['POST'])
def login():
    try:
        data = request.json
        email = data.get('email', '')
        password = data.get('password', '')

        if email == ADMIN_EMAIL and password == ADMIN_PASSWORD:
            token = jwt.encode(
                {
                    'email': email,
                    'role': 'admin',
                    'exp': datetime.now(timezone.utc) + timedelta(hours=24)
                },
                JWT_SECRET,
                algorithm='HS256'
            )
            return jsonify({
                'status': 'success',
                'token': token,
                'user': {'email': email, 'role': 'admin'}
            }), 200
        else:
            return jsonify({'error': 'Credenciales inválidas'}), 401
    except Exception as e:
        logger.error(f'Login error: {e}')
        return jsonify({'error': str(e)}), 500


def get_next_id(collection_name):
    """Auto-increment ID generator using a counters collection."""
    result = col_counters.find_one_and_update(
        {'_id': collection_name},
        {'$inc': {'seq': 1}},
        upsert=True,
        return_document=True
    )
    return result['seq']


# ===================== API Clientes =====================

@app.route('/api/clients', methods=['GET'])
def get_clients():
    try:
        clients = list(col_clients.find({}, {'_id': 0}).sort('id', -1))
        return jsonify(clients)
    except Exception as e:
        logger.error(f"Error getting clients: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/clients', methods=['POST'])
def add_client():
    try:
        data = request.json
        logger.info(f"Adding client: {data}")
        client = {
            'id': get_next_id('clients'),
            'nombre': data['nombre'],
            'apellido': data['apellido'],
            'telefono': data.get('telefono'),
            'correo': data.get('correo'),
            'nota': data.get('nota')
        }
        col_clients.insert_one(client)
        return jsonify({"status": "success"}), 201
    except Exception as e:
        logger.error(f"Error adding client: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/clients/<int:client_id>', methods=['DELETE'])
def delete_client(client_id):
    try:
        # Delete payments for all client loans
        loan_ids = [l['id'] for l in col_loans.find({'client_id': client_id}, {'id': 1, '_id': 0})]
        if loan_ids:
            col_payments.delete_many({'loan_id': {'$in': loan_ids}})
        # Delete loans
        col_loans.delete_many({'client_id': client_id})
        # Delete client
        col_clients.delete_one({'id': client_id})
        return jsonify({"status": "success"})
    except Exception as e:
        logger.error(f"Error deleting client: {e}")
        return jsonify({"error": str(e)}), 500


# ===================== API Socios (Partners) =====================

@app.route('/api/partners', methods=['GET'])
def get_partners():
    try:
        partners = list(col_partners.find({}, {'_id': 0}).sort('id', -1))
        return jsonify(partners)
    except Exception as e:
        logger.error(f"Error getting partners: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/partners', methods=['POST'])
def add_partner():
    try:
        data = request.json
        logger.info(f"Adding partner: {data}")
        partner = {
            'id': get_next_id('partners'),
            'nombre': data['nombre'],
            'nota': data.get('nota')
        }
        col_partners.insert_one(partner)
        return jsonify({"status": "success"}), 201
    except Exception as e:
        logger.error(f"Error adding partner: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/partners/<int:partner_id>', methods=['DELETE'])
def delete_partner(partner_id):
    try:
        col_partners.delete_one({'id': partner_id})
        return jsonify({"status": "success"})
    except Exception as e:
        logger.error(f"Error deleting partner: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/partners/<int:partner_id>/loans', methods=['GET'])
def get_partner_loans(partner_id):
    try:
        # Aggregate loans with client info
        pipeline = [
            {'$match': {'partner_id': partner_id}},
            {'$lookup': {
                'from': 'clients',
                'localField': 'client_id',
                'foreignField': 'id',
                'as': 'client'
            }},
            {'$unwind': '$client'},
            {'$sort': {'id': -1}}
        ]
        loans = list(col_loans.aggregate(pipeline))
        formatted_loans = []
        for loan in loans:
            formatted_loans.append({
                "id": loan["id"],
                "clientId": loan["client_id"],
                "clientName": f"{loan['client']['nombre']} {loan['client']['apellido']}",
                "fecha": loan["fecha"],
                "fechaFin": loan.get("fecha_fin"),
                "monto": loan["monto"],
                "porcentaje": loan["porcentaje"],
                "total": loan["total"],
                "status": loan["status"],
                "parentId": loan.get("parent_id"),
                "partnerId": loan.get("partner_id"),
                "partnerPercentage": loan.get("partner_percentage"),
                "partnerCapital": loan.get("partner_capital"),
                "active": loan.get("active", 1)
            })
        return jsonify(formatted_loans)
    except Exception as e:
        logger.error(f"Error getting partner loans: {e}")
        return jsonify({"error": str(e)}), 500


# ===================== API Préstamos =====================

@app.route('/api/clients/<int:client_id>/loans', methods=['GET'])
def get_loans(client_id):
    try:
        loans = list(col_loans.find({'client_id': client_id}, {'_id': 0}).sort('id', -1))
        formatted_loans = []
        for loan in loans:
            formatted_loans.append({
                "id": loan["id"],
                "clientId": loan["client_id"],
                "fecha": loan["fecha"],
                "fechaFin": loan.get("fecha_fin"),
                "monto": loan["monto"],
                "porcentaje": loan["porcentaje"],
                "total": loan["total"],
                "status": loan["status"],
                "parentId": loan.get("parent_id"),
                "partnerId": loan.get("partner_id"),
                "partnerPercentage": loan.get("partner_percentage"),
                "partnerCapital": loan.get("partner_capital"),
                "active": loan.get("active", 1)
            })
        return jsonify(formatted_loans)
    except Exception as e:
        logger.error(f"Error getting loans: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/loans', methods=['POST'])
def add_loan():
    try:
        data = request.json
        logger.info(f"Adding loan: {data}")
        loan = {
            'id': get_next_id('loans'),
            'client_id': data['clientId'],
            'fecha': data['fecha'],
            'fecha_fin': data.get('fechaFin'),
            'monto': data['monto'],
            'porcentaje': data['porcentaje'],
            'total': data['total'],
            'status': data['status'],
            'parent_id': data.get('parentId'),
            'partner_id': data.get('partnerId'),
            'partner_percentage': data.get('partnerPercentage'),
            'partner_capital': data.get('partnerCapital'),
            'active': 1
        }
        col_loans.insert_one(loan)
        return jsonify({"status": "success"}), 201
    except Exception as e:
        logger.error(f"Error adding loan: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/loans/<int:loan_id>', methods=['DELETE'])
def delete_loan(loan_id):
    try:
        col_payments.delete_many({'loan_id': loan_id})
        col_loans.delete_one({'id': loan_id})
        return jsonify({"status": "success"})
    except Exception as e:
        logger.error(f"Error deleting loan: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/loans/<int:loan_id>', methods=['PUT'])
def update_loan(loan_id):
    try:
        data = request.json
        logger.info(f"Updating loan {loan_id}: {data}")

        # Map frontend camelCase keys to DB snake_case keys
        key_map = {
            'clientId': 'client_id',
            'fechaFin': 'fecha_fin',
            'parentId': 'parent_id',
            'partnerId': 'partner_id',
            'partnerPercentage': 'partner_percentage',
            'partnerCapital': 'partner_capital'
        }
        allowed_fields = ['client_id', 'fecha', 'fecha_fin', 'monto', 'porcentaje',
                          'total', 'status', 'parent_id', 'partner_id',
                          'partner_percentage', 'partner_capital', 'active']

        update_fields = {}
        for key, value in data.items():
            db_key = key_map.get(key, key)
            if db_key in allowed_fields:
                update_fields[db_key] = value

        if not update_fields:
            return jsonify({"error": "No valid fields to update"}), 400

        col_loans.update_one({'id': loan_id}, {'$set': update_fields})
        return jsonify({"status": "success"})
    except Exception as e:
        logger.error(f"Error updating loan: {e}")
        return jsonify({"error": str(e)}), 500


# ===================== API Pagos =====================

@app.route('/api/loans/<int:loan_id>/payments', methods=['GET'])
def get_payments(loan_id):
    try:
        payments = list(col_payments.find({'loan_id': loan_id}, {'_id': 0}).sort('id', 1))
        return jsonify(payments)
    except Exception as e:
        logger.error(f"Error getting payments: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/loans/<int:loan_id>/payments', methods=['POST'])
def add_payment(loan_id):
    try:
        data = request.json
        payment = {
            'id': get_next_id('payments'),
            'loan_id': loan_id,
            'monto': data['monto'],
            'fecha': data['fecha'],
            'saldoAnterior': data.get('saldoAnterior'),
            'interes': data.get('interes'),
            'saldoNuevo': data.get('saldoNuevo'),
            'nota': data.get('nota')
        }
        col_payments.insert_one(payment)

        # Check if loan is fully paid
        loan = col_loans.find_one({'id': loan_id}, {'total': 1, '_id': 0})
        if loan:
            pipeline = [
                {'$match': {'loan_id': loan_id}},
                {'$group': {'_id': None, 'total_paid': {'$sum': '$monto'}}}
            ]
            result = list(col_payments.aggregate(pipeline))
            total_paid = result[0]['total_paid'] if result else 0
            if total_paid >= loan['total']:
                col_loans.update_one({'id': loan_id}, {'$set': {'status': 'pagado'}})

        return jsonify({"status": "success"}), 201
    except Exception as e:
        logger.error(f"Error adding payment: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/payments/<int:payment_id>', methods=['DELETE'])
def delete_payment(payment_id):
    try:
        payment = col_payments.find_one({'id': payment_id})
        if payment:
            loan_id = payment['loan_id']

            # Revert capitalized interest if present
            interest_added = payment.get('interes', 0)
            if interest_added and interest_added > 0:
                col_loans.update_one(
                    {'id': loan_id}, 
                    {'$inc': {'total': -interest_added}}
                )

            col_payments.delete_one({'id': payment_id})

            # Re-check loan status
            loan = col_loans.find_one({'id': loan_id}, {'total': 1, '_id': 0})
            if loan:
                pipeline = [
                    {'$match': {'loan_id': loan_id}},
                    {'$group': {'_id': None, 'total_paid': {'$sum': '$monto'}}}
                ]
                result = list(col_payments.aggregate(pipeline))
                total_paid = result[0]['total_paid'] if result else 0
                if total_paid < loan['total']:
                    col_loans.update_one({'id': loan_id}, {'$set': {'status': 'pendiente'}})
        return jsonify({"status": "success"})
    except Exception as e:
        logger.error(f"Error deleting payment: {e}")
        return jsonify({"error": str(e)}), 500


# ===================== Report Generation =====================

@app.route('/api/clients/<int:client_id>/report/pdf', methods=['GET'])
def generate_client_pdf(client_id):
    try:
        client = col_clients.find_one({'id': client_id}, {'_id': 0})
        if not client:
            return jsonify({"error": "Client not found"}), 404

        loans = list(col_loans.find({'client_id': client_id}, {'_id': 0}).sort('id', 1))

        # Generate PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        elements.append(Paragraph(f"Reporte de Cliente: {client['nombre']} {client['apellido']}", styles['Title']))
        elements.append(Paragraph(f"Teléfono: {client.get('telefono', 'N/A')} | Correo: {client.get('correo', 'N/A')}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Table Data
        data = [['ID', 'Fecha', 'Monto', 'Total', 'Estado', 'Activo']]
        total_monto = 0
        total_deuda = 0

        for loan in loans:
            m = loan['monto']
            t = loan['total']
            total_monto += m
            total_deuda += t

            data.append([
                str(loan['id']),
                loan['fecha'],
                f"${m:,.2f}",
                f"${t:,.2f}",
                loan['status'],
                'Sí' if loan.get('active', 1) else 'No'
            ])

        # Totals Row
        data.append([
            'TOTALES', '', f"${total_monto:,.2f}", f"${total_deuda:,.2f}", '', ''
        ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"reporte_{client['nombre']}_{client['apellido']}.pdf",
            mimetype='application/pdf'
        )

    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/clients/<int:client_id>/report/excel', methods=['GET'])
def generate_client_excel(client_id):
    try:
        client = col_clients.find_one({'id': client_id}, {'_id': 0})
        if not client:
            return jsonify({"error": "Client not found"}), 404

        loans = list(col_loans.find({'client_id': client_id}, {'_id': 0}).sort('id', 1))

        # Generate Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Préstamos"

        # Header Info
        ws['A1'] = f"Reporte de Cliente: {client['nombre']} {client['apellido']}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')

        ws['A2'] = f"Teléfono: {client.get('telefono', 'N/A')}"
        ws['A3'] = f"Correo: {client.get('correo', 'N/A')}"

        # Table Headers
        headers = ['ID', 'Fecha', 'Monto', 'Total', 'Estado', 'Activo']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data
        total_monto = 0
        total_deuda = 0
        current_row = 6

        for loan in loans:
            m = loan['monto']
            t = loan['total']
            total_monto += m
            total_deuda += t

            ws.cell(row=current_row, column=1, value=loan['id'])
            ws.cell(row=current_row, column=2, value=loan['fecha'])
            ws.cell(row=current_row, column=3, value=m).number_format = '$#,##0.00'
            ws.cell(row=current_row, column=4, value=t).number_format = '$#,##0.00'
            ws.cell(row=current_row, column=5, value=loan['status'])
            ws.cell(row=current_row, column=6, value='Sí' if loan.get('active', 1) else 'No')
            current_row += 1

        # Totals Row
        ws.cell(row=current_row, column=1, value="TOTALES")
        ws.cell(row=current_row, column=3, value=total_monto).number_format = '$#,##0.00'
        ws.cell(row=current_row, column=4, value=total_deuda).number_format = '$#,##0.00'

        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"reporte_{client['nombre']}_{client['apellido']}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Error generating Excel: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/reports/all-loans/pdf', methods=['GET'])
def generate_general_pdf():
    try:
        # Aggregate loans with client info
        pipeline = [
            {'$lookup': {
                'from': 'clients',
                'localField': 'client_id',
                'foreignField': 'id',
                'as': 'client'
            }},
            {'$unwind': '$client'},
            {'$sort': {'id': -1}}
        ]
        loans = list(col_loans.aggregate(pipeline))

        # Generate PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        elements.append(Paragraph("Reporte General de Préstamos", styles['Title']))
        elements.append(Spacer(1, 12))

        # Table Data
        data = [['ID', 'Cliente', 'Monto', 'Total', 'Estado']]
        total_monto = 0
        total_deuda = 0

        for loan in loans:
            m = loan['monto']
            t = loan['total']
            total_monto += m
            total_deuda += t

            data.append([
                str(loan['id']),
                f"{loan['client']['nombre']} {loan['client']['apellido']}",
                f"${m:,.2f}",
                f"${t:,.2f}",
                loan['status']
            ])

        # Totals Row
        data.append([
            'TOTALES', '', f"${total_monto:,.2f}", f"${total_deuda:,.2f}", ''
        ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name="reporte_general_prestamos.pdf",
            mimetype='application/pdf'
        )

    except Exception as e:
        logger.error(f"Error generating General PDF: {e}")
        return jsonify({"error": str(e)}), 500

# ===================== Import SQLite DB =====================

@app.route('/api/import-db', methods=['POST'])
def import_db():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files['file']
        if not file.filename.endswith('.db'):
            return jsonify({"error": "File must be a .db SQLite file"}), 400

        # Save to temp file
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
        file.save(tmp.name)
        tmp.close()

        try:
            conn = sqlite3.connect(tmp.name)
            conn.row_factory = sqlite3.Row

            # Clear existing MongoDB data
            col_clients.delete_many({})
            col_partners.delete_many({})
            col_loans.delete_many({})
            col_payments.delete_many({})
            col_counters.delete_many({})

            counts = {}

            # --- Import Clients ---
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM clients')
            rows = [dict(r) for r in cursor.fetchall()]
            if rows:
                col_clients.insert_many(rows)
            counts['clients'] = len(rows)

            # --- Import Partners ---
            try:
                cursor.execute('SELECT * FROM partners')
                rows = [dict(r) for r in cursor.fetchall()]
                if rows:
                    col_partners.insert_many(rows)
                counts['partners'] = len(rows)
            except Exception:
                counts['partners'] = 0

            # --- Import Loans ---
            cursor.execute('SELECT * FROM loans')
            raw_loans = [dict(r) for r in cursor.fetchall()]
            loans_to_insert = []
            for loan in raw_loans:
                # Normalize column names (handle both clientId and client_id)
                doc = {
                    'id': loan.get('id'),
                    'client_id': loan.get('client_id') or loan.get('clientId'),
                    'fecha': loan.get('fecha'),
                    'fecha_fin': loan.get('fecha_fin') or loan.get('fechaFin'),
                    'monto': loan.get('monto'),
                    'porcentaje': loan.get('porcentaje'),
                    'total': loan.get('total'),
                    'status': loan.get('status'),
                    'parent_id': loan.get('parent_id') or loan.get('parentId'),
                    'partner_id': loan.get('partner_id') or loan.get('partnerId'),
                    'partner_percentage': loan.get('partner_percentage') or loan.get('partnerPercentage'),
                    'partner_capital': loan.get('partner_capital') or loan.get('partnerCapital'),
                    'active': loan.get('active', 1)
                }
                loans_to_insert.append(doc)
            if loans_to_insert:
                col_loans.insert_many(loans_to_insert)
            counts['loans'] = len(loans_to_insert)

            # --- Import Payments ---
            cursor.execute('SELECT * FROM payments')
            rows = [dict(r) for r in cursor.fetchall()]
            if rows:
                col_payments.insert_many(rows)
            counts['payments'] = len(rows)

            conn.close()

            # Reset auto-increment counters to max IDs
            for name, collection in [('clients', col_clients), ('partners', col_partners),
                                      ('loans', col_loans), ('payments', col_payments)]:
                pipeline = [{'$group': {'_id': None, 'max_id': {'$max': '$id'}}}]
                result = list(collection.aggregate(pipeline))
                max_id = result[0]['max_id'] if result else 0
                col_counters.update_one(
                    {'_id': name},
                    {'$set': {'seq': max_id or 0}},
                    upsert=True
                )

            logger.info(f"Import completed: {counts}")
            return jsonify({"status": "success", "imported": counts}), 200

        finally:
            os.unlink(tmp.name)

    except Exception as e:
        logger.error(f"Error importing DB: {e}")
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5000)
